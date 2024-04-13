from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponseRedirect
import pickle
from manager.models import department_description
from manager.models import anouncement
from manager.models import CDC_FD
from manager.models import CDC_HD
from manager.models import Elective_FD
from manager.models import Elective_HD
from manager.models import WILP
from manager.models import Faculty_List
from openpyxl import Workbook
from openpyxl import load_workbook
from .forms import classformuser
from .forms import facultyform1user
from .forms import facultyform2user
from .forms import facultyform3user
from .forms import Electiveform_FDuser
from .forms import Electiveform_HDuser
from .forms import Semform
import datetime
from django.forms import formset_factory
from pathlib import Path
import pandas as pd
import numpy as np
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.csrf import csrf_exempt
import os
from localStoragePy import localStoragePy
from django.forms.models import model_to_dict
from django.db.models import Q




class MyFormsetFactory:
    def __init__(self, form, phd_initial=None, faculty_initial=None, extra=1):
        self.formset_factory = formset_factory(form, extra=extra)
        self.phd_initial = phd_initial
        self.faculty_initial = faculty_initial

    def __call__(self, *args, **kwargs):
        formset = self.formset_factory(*args, **kwargs)
        self._set_initial_data(formset)
        return formset

    def _set_initial_data(self, formset):
        for index, form in enumerate(formset.forms):
            if self.phd_initial:
                form.fields['PHD'].initial = self.phd_initial[index] if index < len(self.phd_initial) else None
            if self.faculty_initial:
               form.fields['Faculty'].initial = self.faculty_initial[index] if index < len(self.faculty_initial) else None





def logout_user(request):
    logout(request)
    return  HttpResponseRedirect('login_user?logout=True')

localStorage = localStoragePy('user', 'text')
def login_user(request):
    
    
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if department_description.objects.get(Department_HOD=request.user).Lock==False:
                return redirect('home')
            else:
                messages.success(
                request, "Administrator has locked the website.")
                return redirect('login')
        

        else:
            messages.success(
                request, "Username or Password might be incorrect.Please try again or contact the administrator.")
            return redirect('login')

    else:
        
        return render(request, 'authentication/login.html', {})


def home(request):
   
    announcement_data = anouncement.objects.all()
   

    return render(request, 'homepage/home.html', {'announcement': announcement_data})


def previousrecord(request):
    try:
        link=department_description.objects.get(Department_HOD=request.user).Previous_records
    except TypeError:
        return HttpResponseRedirect('login_user')
    return render(request, 'homepage/previousrecord.html', {'link':link})


def upload(request):
    try:
        Department=department_description.objects.get(Department_HOD=request.user)
    except TypeError:
        return HttpResponseRedirect('login_user')
    academic_year=str(str(datetime.date.today().year)+"-"+str(int(datetime.date.today().year)+1))
    current_sem=department_description.objects.get(Department_HOD=request.user).Upcoming_Sem
    context = {"academic_year":academic_year,"Department":Department}
    if request.method == "POST":
        uploaded_file = request.FILES['document']
        fs = FileSystemStorage(location=str(academic_year)+"/"+current_sem+"/"+str(Department))
        name = fs.save(uploaded_file.name, uploaded_file)
        context['url'] = fs.url(str(academic_year)+"/"+current_sem+"/"+str(Department)+"/"+str(name))
    return render(request, 'homepage/upload.html', context)


def filldata(request):
    return render(request, 'homepage/filldata.html', {})

@csrf_exempt
def choose_new_table(request):
    filecount=0
    academic_year_folder=str(str(datetime.date.today().year)+"-"+str(int(datetime.date.today().year)+1))
    try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
    except TypeError:
        return HttpResponseRedirect('login_user')
    current_sem=department_description.objects.get(Department_HOD=request.user).Upcoming_Sem
    path=str(academic_year_folder)+"/"+current_sem+"/"+str(Department_name)
   
    if not os.path.exists(path):
        os.makedirs(path)

    try:
        for filename in os.listdir(path):
            if filename.endswith('xlsx'): 
                filecount=filecount+1
          
        attempts_remaining=3-filecount
   

        if attempts_remaining<=0:
            df=""
            message="Maximum submissions perfomed."
            tablePresent=False

        else:
            df=pd.read_excel("Pickles/Courses for Course Load Submission "+str(Department_name)+".xlsx",sheet_name=str(Department_name))
            df=df.drop([0])
            df=df.fillna("")
            message=""
            tablePresent=True
            df.reset_index(drop=True, inplace=True)
            df=df.to_html(index=False)
    except FileNotFoundError:
        df=""
        message="Table not created yet."
        tablePresent=False
        attempts_remaining=3-filecount
    
    if request.GET.get("elective_hd"):
            return Elective_HD_list(request=request)
    if request.GET.get("elective_fd"):
            return Elective_FD_list(request=request)
    if request.method=="POST":

        form=Semform(request.POST or None)
        if form.is_valid():
           
        
            file=load_workbook("Pickles/"+"Courses for Course Load Submission "+str(Department_name)+".xlsx")
            sheet=file.get_sheet_by_name(str(Department_name))
            sheet["J2"]=form.cleaned_data["Sidenote"]
            file_path=Path(path+"/Courses for Course Load Submission "+str(current_sem)+" "+str(Department_name)+".xlsx")
           
            if file_path.exists():
               
                file_old=load_workbook(file_path)
                file_old_path=Path(path+"/Courses for Course Load Submission "+str(current_sem)+" "+str(Department_name)+" Old_1"+".xlsx")
                file_older_path=Path(path+"/Courses for Course Load Submission "+str(current_sem)+" "+str(Department_name)+" Old_2"+".xlsx")
                if file_old_path.exists():
                    file_older=load_workbook(file_old_path)
                    file_older.save(file_older_path)
                else:
                    file_old.save(file_old_path)
            
            file.save(path+"/Courses for Course Load Submission "+str(current_sem)+" "+str(Department_name)+".xlsx")
    else:
        form=Semform()
    
    
    return render(request, 'homepage/choose_new_table.html', {"message":message,"table":df,"bool":tablePresent,"form":form,"attempts":attempts_remaining,"sem":current_sem})

 
def CDC_FD_list(request):
    try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
        Semester=department_description.objects.get(Department_HOD=request.user).Upcoming_Sem
       
    except TypeError:
        return HttpResponseRedirect('login_user')
    CDC_name=[]
    CDC_department = CDC_FD.objects.filter(Q(CDC_Department=Department_name) & Q(Upcoming_Sem_FD=Semester))
    # CDC_name.extend( CDC_FD.objects.filter(Q(CDC_Department=Department_name) & Q(Upcoming_Sem_FD=Semester)).get().CDC_name)
    
    try:
        with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'rb') as f:
           
            cdc_department = pickle.load(f)
    except FileNotFoundError:
        cdc_department = []
    
    # Filter CDC_FD objects based on cdc_department entries
    matched_CDC_department = []

    for course_title in cdc_department:
        matched_CDC_department.extend(CDC_FD.objects.filter(Q(CDC_Department=Department_name) & Q(Upcoming_Sem_FD=Semester) & Q(CDC_name=course_title)))
   
    return render(request, 'homepage/CDC_FD_list.html', {"CDC_List": CDC_department,"match_list":matched_CDC_department,"CDC_name_list":CDC_name})

def CDC_HD_list(request):
    try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
        Semester=department_description.objects.get(Department_HOD=request.user).Upcoming_Sem
    except TypeError:
        return HttpResponseRedirect('login_user')
    CDC_department = (CDC_HD.objects.filter(Q(CDC_HD_Department=Department_name) & Q(Upcoming_Sem_HD=Semester)))
    try:
        with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'rb') as f:
            
            cdc_department = pickle.load(f)
    except FileNotFoundError:
        cdc_department = []
   

    # Filter CDC_FD objects based on cdc_department entries
    matched_CDC_department = []
    for course_title in cdc_department:
        matched_CDC_department.extend(CDC_HD.objects.filter(Q(CDC_HD_Department=Department_name) & Q(Upcoming_Sem_HD=Semester) & Q(CDC_HD_name=course_title)))
   
   
    return render(request, 'homepage/CDC_HD_list.html', {"CDC_List": CDC_department,"match_list":matched_CDC_department})

def WILP_list(request):
    try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
        
    except TypeError:
        return HttpResponseRedirect('login_user')
    WILP_department = (WILP.objects.filter(Q(WILP_Department=Department_name) ))
    try:
        with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'rb') as f:
           
            cdc_department = pickle.load(f)
    except FileNotFoundError:
        cdc_department = []
    
    # Filter CDC_FD objects based on cdc_department entries
    matched_CDC_department = []
    for course_title in cdc_department:
        matched_CDC_department.extend(WILP.objects.filter(Q(WILP_Department=Department_name)  & Q(WILP_name=course_title)))
   
    return render(request, 'homepage/CDC_HD_list.html', {"WILP_List": WILP_department,"match_list":matched_CDC_department})
def previewForm(request):
    global course_id


    
    if request.method == 'GET':
           course_id = request.GET.get('data')
           localStorage.setItem('course_id',course_id)

    current_sem=department_description.objects.get(Department_HOD=request.user).Upcoming_Sem
    academic_year_folder=str(department_description.objects.get(Department_HOD=request.user).Academic_year)
    Department_name = department_description.objects.get(Department_HOD=request.user)

    path=str(str(academic_year_folder)+"/"+current_sem+"/"+str(Department_name))
   
    path1=('Pickles/Courses for Course Load Submission '+str(Department_name)+'.xlsx')

    if Path(path1).exists():
        try:
            database=pd.read_excel(path1,sheet_name=str(Department_name))
        except (ValueError,UnboundLocalError,FileNotFoundError):
             database=pd.read_excel("cache.xlsx")

    else:
        for filename in os.listdir(path):
            if filename.endswith('xlsx'): 
                filepath=os.path.join(path, filename)
       
        try:
            database=pd.read_excel(filepath,sheet_name=str(Department_name))
        
        except (ValueError,UnboundLocalError,FileNotFoundError):
            database=pd.read_excel("cache.xlsx")
            
       

    
   
    try:
        database=database.drop([0])
        database=database.drop(database.columns[7],axis=1)
        database=database.drop(database.columns[8],axis=1)
    except (IndexError,KeyError):
        print()
    
    try:
        row_number=database[database.eq(course_id).any(axis=1)].index.to_numpy()
        row_count=0
        for i in range(row_number[0],len(database[database.columns[1]])-1):
            if database.iloc[i][0] is not np.nan:
                break
            row_count=row_count+1  


        section_list=[]

        for i in range(0,row_count+1):
            section_list.append(database.iloc[row_number[0]+i][4])
       
        section_list_numpy=np.array(section_list)
        Number_of_Lectures=np.count_nonzero(section_list_numpy=="L")+1
        Number_of_Tutorials=np.count_nonzero(section_list_numpy=="T")
        Number_of_Labs=np.count_nonzero(section_list_numpy=="P")
        print(section_list)
    except IndexError:
        Number_of_Lectures=0
        Number_of_Labs=0
    
        Number_of_Tutorials=0
 
    Lecture_Faculty=[]
    Tutorial_Faculty=[]
    Labarotary_Faculty=[]
    try:
        FIC_preview=str(database.iat[int(row_number-1),5])
    except TypeError:
        FIC_preview=""
    
    for i in range(0,Number_of_Lectures):
     if str(database.iat[int(i+row_number-1),6])=="nan":
         Lecture_Faculty.append("No Data")
     else:    
        Lecture_Faculty.append(str(database.iat[int(i+row_number-1),6]))
    for i in range(0,Number_of_Labs):
      if str(database.iat[int(i+row_number+Number_of_Tutorials+Number_of_Lectures-1),6])=="nan":
         Labarotary_Faculty.append("No Data")
      else:   
         Labarotary_Faculty.append(str(database.iat[int(i+row_number+Number_of_Tutorials+Number_of_Lectures-1),6]))
    for i in range(0,Number_of_Tutorials):
        if str(database.iat[int(i+row_number+Number_of_Lectures),6])=="nan":
         Tutorial_Faculty.append("No Data")
        else: 
         Tutorial_Faculty.append(str(database.iat[int(i+row_number+Number_of_Lectures-1),6]))
     

    
    try:
        
        course_title=CDC_FD.objects.filter(CDC_ID=course_id).first().CDC_name  
        print() 
        localStorage.setItem("course_title",course_title)
    except (ObjectDoesNotExist, AttributeError):
        print()
        try:
            course_title=CDC_HD.objects.filter(CDC_HD_ID=course_id).first().CDC_HD_name  
            localStorage.setItem("course_title",course_title) 
        except (ObjectDoesNotExist, AttributeError):
            print("")
            try:
                course_title=Elective_HD.objects.filter(Elective_HD_ID=course_id).first().Elective_HD_name
                localStorage.setItem("course_title",course_title)   
            except (ObjectDoesNotExist, AttributeError):
                print()
                try:
                    course_title=Elective_FD.objects.filter(Elective_ID=course_id).first().Elective_name
                    localStorage.setItem("course_title",course_title)   
                except (ObjectDoesNotExist, AttributeError):
                    print()
    data = {
    'FIC': FIC_preview,
    'Number_of_Lectures': Number_of_Lectures,
    'Number_of_Labs': Number_of_Labs,
    'Number_of_Tutorials': Number_of_Tutorials,
    'Lecture_Faculty': Lecture_Faculty,
    'Labaratory_Faculty': Labarotary_Faculty,
    'Tutorial_Faculty': Tutorial_Faculty
    }


    try:    
                    f=open("Pickles/"+str(Department_name)+'_form.pkl','wb')
    except FileNotFoundError:
                    open("Pickles/"+str(Department_name)+'_form.pkl','a')
                    f=open("Pickles/"+str(Department_name)+'_form.pkl','wb')


    (pickle.dump(data,f))

        
    if request.method=="POST":
        
        create_file(request=request,FIC_name=FIC_preview,Lecture=Number_of_Lectures,Lab=Number_of_Labs,Tutorial=Number_of_Tutorials,Faculty_Lec=Lecture_Faculty,Faculty_Tut=Tutorial_Faculty,Faculty_Lab=Labarotary_Faculty)
        return  HttpResponseRedirect('choose_new_table')
    return render(request,'homepage/previewForm.html',{'Lectures':Number_of_Lectures,'Tutorial':Number_of_Tutorials,'Labs':Number_of_Labs,"CDC_name":course_title,"Lab_Faculty":Labarotary_Faculty,"Tut_Faculty":Tutorial_Faculty,"Lec_Faculty":Lecture_Faculty})





def form_CDC(request):
    global FIC
    global course_title
    try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
    except TypeError:
        return HttpResponseRedirect('login_user')
           
    course_title = request.GET.get('data')
    label=request.GET.get('label')
    localStorage.setItem('label',label)
    if label=='No':
        formclass=classformuser(user=request.user,initial_val={'Lectures':0,'Tutorial':0,'Labs':0})
    elif label=='Modify':
        try:
            with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                data = pickle.load(f)
        except FileNotFoundError:
            open("Pickles/"+str(Department_name)+'_form.pkl','a')
            with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                data = pickle.load(f)   
        formclass=classformuser(user=request.user,initial_val={'Lectures':int(data['Number_of_Lectures']),'Tutorial':int(data['Number_of_Tutorials']),'Labs':int(data['Number_of_Labs'])})
   
        
        

   
    if request.method == 'POST':
        form = formclass(request.POST) 

        if form.is_valid():
                FIC=form.cleaned_data["FIC"]
                localStorage.setItem('cdc_FIC',FIC)
                request.session['Tuts']=form.cleaned_data["Tutorials"]
                request.session['Lec']=form.cleaned_data["Lectures"]
                request.session['Lab']=form.cleaned_data["Labs"]
                
                
                
        return form_faculty_lec(request=request)
                
       
                           
    
    else:
         
         form = formclass()
    

    return render(request, "homepage/form_CDC.html", context={'form': form,"cdc_name":course_title})



def form_faculty_lec(request):
 try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
 except TypeError:
        return HttpResponseRedirect('login_user')
 
 global Lab_number
 global Lecture_Number
 global Tutorial_number
 global Lec_Faculty
 Lec_Faculty=[]
 Lab_number=request.session["Lab"]
 Lecture_Number=request.session["Lec"]
 Tutorial_number=request.session["Tuts"] 
 localStorage.setItem('Lab_number',Lab_number)
 localStorage.setItem('Tutorial_number',Tutorial_number)
 localStorage.setItem('Lecture_number',Lecture_Number)
 label=localStorage.getItem('label')
 if label == 'No':
    facultyform1=facultyform1user(user=request.user)
    Lectureformset=formset_factory(facultyform1,extra=Lecture_Number)
 elif label=='Modify':
      try:
            with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                data = pickle.load(f)
      except FileNotFoundError:
            open("Pickles/"+str(Department_name)+'_form.pkl','a')
            with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                data = pickle.load(f)
      facultyform1=facultyform1user(user=request.user)
      faculty_names = data['Lecture_Faculty']
      faculty_names_split = [
    entry.strip() for names in faculty_names for entry in names.split('/')

]

        # faculty_names_prepended now contains the modified strings with "Prepended Data" added to each entry
     
      Lectureformset=MyFormsetFactory(facultyform1,phd_initial=faculty_names_split,faculty_initial=faculty_names_split,extra=Lecture_Number)

 



    
 if request.method=="POST":
       
 


        form_lec=Lectureformset(request.POST or None)
       
        if form_lec.is_valid(): 
            try:
             for form in form_lec:
                faculty_names = '/'.join(faculty.first_name for faculty in form.cleaned_data['Faculty'])
                phd_names= '/'.join(faculty.first_name for faculty in form.cleaned_data['PHD'])
                overall_names=faculty_names+'/'+phd_names
                Lec_Faculty.append(overall_names)
            except KeyError:
                return HttpResponseRedirect('form_Faculty_Lec') 
           
            return  HttpResponseRedirect('form_Faculty_Tut')
                
            

 return render(request, 'homepage/facultyForm_Lec.html', {'Lectureformset': Lectureformset})






def form_faculty_tut(request):
        try:
            Department_name = department_description.objects.get(Department_HOD=request.user)
        except TypeError:
            return HttpResponseRedirect('login_user')
        
        global Tut_Faculty
        Tut_Faculty=[]
        
       
        Tutorial_number= int(localStorage.getItem('Tutorial_number'))
        
        label=localStorage.getItem('label')
        if label == 'No':
            facultyform2=facultyform2user(user=request.user)
            Tutformset=formset_factory(facultyform2,extra=int(Tutorial_number))
        elif label=='Modify':
            try:
                with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                    data = pickle.load(f)
            except FileNotFoundError:
                    open("Pickles/"+str(Department_name)+'_form.pkl','a')
                    with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                        data = pickle.load(f)
            facultyform2=facultyform2user(user=request.user)
            faculty_names = data['Tutorial_Faculty']
            Tutformset=MyFormsetFactory(facultyform2,phd_initial=data['Tutorial_Faculty'],faculty_initial=faculty_names,extra=int(Tutorial_number))

    
        if request.method=="POST":
            tag=request.POST.get('clear')
            if tag=='clear':
                return HttpResponseRedirect('form_Faculty_Tut')
            form_tut=Tutformset(request.POST or None)

            if form_tut.is_valid():
                try:
                 for form in form_tut:
                     faculty_names = '/'.join(faculty.first_name for faculty in form.cleaned_data['Faculty'])
                     phd_names= '/'.join(faculty.first_name for faculty in form.cleaned_data['PHD'])
                     overall_names=faculty_names+'/'+phd_names
                     Tut_Faculty.append(overall_names)
                except KeyError:
                    return HttpResponseRedirect('form_Faculty_Tut') 

                return  HttpResponseRedirect('form_Faculty_Lab')
        return render(request, 'homepage/facultyForm_Tut.html', {"Tutformset":Tutformset})



def form_faculty_lab(request):  
         try:
            Department_name = department_description.objects.get(Department_HOD=request.user)
         except TypeError:
            return HttpResponseRedirect('login_user')
        
      
         submitted=False
         Lab_Faculty=[]
         Department_name = department_description.objects.get(Department_HOD=request.user)
         Lab_number= int(localStorage.getItem('Lab_number'))
         
         label=localStorage.getItem('label')
         if label == 'No':
            facultyform3=facultyform3user(user=request.user)
            Labformset=formset_factory(facultyform3,extra=int(Lab_number))
         elif label=='Modify':
            try:
                with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                    data = pickle.load(f)
            except FileNotFoundError:
                    open("Pickles/"+str(Department_name)+'_form.pkl','a')
                    with open("Pickles/"+str(Department_name)+'_form.pkl', 'rb') as f:
                        data = pickle.load(f)
            facultyform3=facultyform3user(user=request.user)
            faculty_names = data['Labaratory_Faculty']
            Labformset=MyFormsetFactory(facultyform3,phd_initial=data['Labaratory_Faculty'],faculty_initial=faculty_names,extra=int(Lab_number))
         
         
         
         
         if request.method=="POST":
            tag=request.POST.get('clear')
     
            if tag=='clear':
                return HttpResponseRedirect('form_Faculty_Lab') 
            form_lab=Labformset(request.POST or None)
            submitted=True
            if form_lab.is_valid():
                try:
                    for form in form_lab:
                        faculty_names = ','.join(faculty.first_name for faculty in form.cleaned_data['Faculty'])
                        phd_names= ','.join(faculty.first_name for faculty in form.cleaned_data['PHD'])
                    
                        overall_names=faculty_names+','+phd_names
                        Lab_Faculty.append(overall_names)
                    

            # Save updated course_titles list to pickle file
                    if os.path.exists('Pickles/course_titles'+str(Department_name)+'.pickle'):
                     # Load existing course_titles from the pickle file
                        with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'rb') as f:
                            course_titles = pickle.load(f)
                    else:
                     # If the pickle file doesn't exist, create an empty list
                        course_titles = []
                    
                    course_titles.append(course_title)
                 
                    with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'wb') as f:
                        pickle.dump(course_titles, f)      
                        
                except KeyError:
                    return HttpResponseRedirect('form_Faculty_Lab') 
                try:
                    create_file(request=request,FIC_name=FIC,Lecture=Lecture_Number,Tutorial=Tutorial_number,Lab=Lab_number,Faculty_Lab=Lab_Faculty,Faculty_Lec=Lec_Faculty,Faculty_Tut=Tut_Faculty)
                    return  HttpResponseRedirect('choose_new_table')
                except NameError:
                    FIC=localStorage.getItem('cdc_FIC')
                    create_file(request=request,FIC_name=FIC,Lecture=Lecture_Number,Tutorial=Tutorial_number,Lab=Lab_number,Faculty_Lab=Lab_Faculty,Faculty_Lec=Lec_Faculty,Faculty_Tut=Tut_Faculty)
                    return  HttpResponseRedirect('choose_new_table')
                
         return render(request, 'homepage/facultyForm_Lab.html', {"Labformset":Labformset,"submitted":submitted})



       
def create_file(request,FIC_name,Lecture,Tutorial,Lab,Faculty_Lec,Faculty_Lab,Faculty_Tut):
   
    course_id=localStorage.getItem('course_id')
    course_title=localStorage.getItem("course_title")
    Department_name = department_description.objects.get(Department_HOD=request.user)
   

    excel_file=Path("Pickles/"+"Courses for Course Load Submission "+str(Department_name)+".xlsx")

    if excel_file.exists():
       file=load_workbook("Pickles/"+"Courses for Course Load Submission "+str(Department_name)+".xlsx")
       sheet=file.get_sheet_by_name(str(Department_name))
       row_if_present=""
       next_row_entry=""
         
       for i in range(1,sheet.max_row):
           column = 'A'
           row = i
           cell_value = sheet[f'{column}{row}'].value
           if cell_value==course_id:
            row_if_present=row
            print(row_if_present)
            break
       if row_if_present!="":
            for i in range(int(row_if_present)+1,sheet.max_row):
                column = 'A'
                row = i
                cell = sheet[f'{column}{row}']
                if cell.value:
                    next_row_entry=row
                    break
            if next_row_entry=="":
                next_row_entry=sheet.max_row
            if row_if_present==1:
                sheet.delete_rows(row_if_present,int(next_row_entry-row_if_present))
            else:
                sheet.delete_rows(row_if_present,int(next_row_entry-row_if_present+1))
           
       last_row=sheet.max_row
       
     
       for i in range(0,Lecture):
           if Lecture==0:
               break
           else:
            try:
                sheet.cell(row=last_row+i+1, column=5).value="L"
                sheet.cell(row=last_row+i+1, column=4).value=i+1
                sheet.cell(row=last_row+i+1, column=7).value=str(Faculty_Lec[i])
                sheet.cell(row=last_row+1,column=6).value=str(FIC_name)
                sheet.cell(row=last_row+1,column=1).value=str(course_id)
                sheet.cell(row=last_row+1,column=2).value=str(course_title)
            except IndexError:
                    break
       for i in range(0,Tutorial):
            try:
                sheet.cell(row=last_row+Lecture+i+1, column=5).value="T"
                sheet.cell(row=Lecture+last_row+i+1, column=4).value=i+1
                sheet.cell(row=last_row+Lecture+i+1, column=7).value=str(Faculty_Tut[i])
            except IndexError:
                break

       for i in range(0,Lab):
         try:
            sheet.cell(row=last_row+Tutorial+Lecture+i+1, column=5).value="P"
            sheet.cell(row=Tutorial+Lecture+last_row+i+1, column=4).value=i+1
            
            sheet.cell(row=last_row+Lecture+Tutorial+i+1, column=7).value=str(Faculty_Lab[i])
         except IndexError:
                break
       print(file)
       file.save("Pickles/"+"Courses for Course Load Submission "+str(Department_name)+".xlsx")

    else:    
   
        Wb_load=Workbook()
        Sheet=Wb_load.create_sheet(title=str(Department_name))
        Sheet['A3']="COURSE NUMBER"
        Sheet['B3']="COURSE TITLE"
        Sheet["D3"]="Section Number"
        Sheet["E3"]="Section"
        Sheet["F3"]="Instructor In Charge"
        Sheet["G3"]="Instructor"
        
       

        for i in range(0,Lecture):
            Sheet.cell(row=4+i, column=4).value=i+1
            Sheet.cell(row=4+i, column=5).value="L"
            Sheet.cell(row=4+i, column=7).value=str(Faculty_Lec[i])
            Sheet['A4']=str(course_id)       
            Sheet['B4']=str(course_title) 
            Sheet["F4"]=str(FIC_name) 
        
        for i in range(0,Tutorial):
          try: 
            Sheet.cell(row=4+Lecture+i, column=5).value="T"
            Sheet.cell(row=4+Lecture+i, column=4).value=i+1
           
            Sheet.cell(row=4+Lecture+i, column=7).value=str(Faculty_Tut[i])
          except IndexError:
            break
        for i in range(0,Lab):
         try:
            Sheet.cell(row=4+Tutorial+Lecture+i, column=5).value="P"
            Sheet.cell(row=4+Tutorial+Lecture+i, column=4).value=i+1
          
            Sheet.cell(row=4+Tutorial+Lecture+i, column=7).value=str(Faculty_Lab[i])
         except IndexError:
            break
        Wb_load.save("Pickles/"+"Courses for Course Load Submission "+str(Department_name)+".xlsx")


def Elective_FD_list(request):
    try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
    except TypeError:
        return HttpResponseRedirect('login_user')
    form_Elective=Electiveform_FDuser(user=request.user)
    try: 
        try:
            with open("Pickles/"+str(Department_name)+'_FD.pkl', 'rb') as f:
                data = pickle.load(f)
        except FileNotFoundError:
            open("Pickles/"+str(Department_name)+'_FD.pkl','a')
            with open("Pickles/"+str(Department_name)+'_FD.pkl', 'rb') as f:
                data = pickle.load(f)
        
        Elective_list= data

        f.close()
        print(Elective_list)
    
        
    except EOFError:
        Elective_list=[]
        form=form_Elective()
    try:
        with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'rb') as f:
            
            cdc_department = pickle.load(f)
    except FileNotFoundError:
        cdc_department = []
    matched_El_department = []
    for course_title in cdc_department:
        matched_El_department.extend(Elective_FD.objects.filter(Q(Elective_Department=Department_name)  & Q(Elective_name=course_title)))
    
    form=form_Elective()
    f.close()
   
    if request.method=="POST":
        form=form_Elective(request.POST or None)
        
        if form.is_valid():
            Elective_list=form.cleaned_data['Elective_ID']
            try:    
                    f=open("Pickles/"+str(Department_name)+'_FD.pkl','wb')
            except FileNotFoundError:
                    open("Pickles/"+str(Department_name)+'_FD.pkl','a')
                    f=open("Pickles/"+str(Department_name)+'_FD.pkl','wb')
            (pickle.dump(Elective_list,f))        
    return render(request,"homepage/Elective_FD_list.html",{"form":form,"elective_list":Elective_list,"match_list":matched_El_department})


def Elective_HD_list(request):
   try:
        Department_name = department_description.objects.get(Department_HOD=request.user)
   except TypeError:
        return HttpResponseRedirect('login_user')
   form_Elective=Electiveform_HDuser(user=request.user)
   try: 
    try:
        with open("Pickles/"+str(Department_name)+'_HD.pkl', 'rb') as f:
            data = pickle.load(f)
    except FileNotFoundError:
            open("Pickles/"+str(Department_name)+'_HD.pkl','a')
            with open("Pickles/"+str(Department_name)+'_HD.pkl', 'rb') as f:
                data = pickle.load(f)
  
    
    
    Elective_list= data
    
   except EOFError:
        Elective_list=[]
        form=form_Elective()
   try:
        with open('Pickles/course_titles'+str(Department_name)+'.pickle', 'rb') as f:
            
            cdc_department = pickle.load(f)
   except FileNotFoundError:
        cdc_department = []
   matched_El_department = []
   for course_title in cdc_department:
        matched_El_department.extend(Elective_HD.objects.filter(Q(Elective_HD_Department=Department_name)  & Q(Elective_HD_name=course_title)))
   form=form_Elective()
   f.close()
   
   if request.method=="POST":
        form=form_Elective(request.POST or None)
        
        if form.is_valid():
               
                Elective_list=form.cleaned_data['Elective_ID']

              
                try:    
                    f=open("Pickles/"+str(Department_name)+'_HD.pkl','wb')
                except FileNotFoundError:
                    open("Pickles/"+str(Department_name)+'_HD.pkl','a')
                    f=open("Pickles/"+str(Department_name)+'_HD.pkl','wb')


                (pickle.dump(Elective_list,f))
  
   return render(request,"homepage/Elective_HD_list.html",{"form":form,"elective_list":Elective_list,"match_list":matched_El_department})


